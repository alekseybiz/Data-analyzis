import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.chrome.service import Service

# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "products.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# URL сайта для поиска
base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта

try:
    # Проходим по всем строкам таблицы
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        product_name = row[0].value  # Название товара из колонки "Товары"

        if not product_name:
            continue

        print(f"Ищем товар: {product_name}")

        # Открываем сайт и выполняем поиск товара
        driver.get(base_url)
        search_box = driver.find_element(By.NAME, "q")  # Найдите селектор для поля поиска
        search_box.clear()
        search_box.send_keys(product_name)
        search_box.send_keys(Keys.RETURN)

        time.sleep(3)  # Ожидание загрузки результатов

        try:
            # Открываем первую ссылку на найденный товар
            product_link = driver.find_element(By.CSS_SELECTOR, "a.product-link")  # Замените на нужный селектор
            product_link.click()

            time.sleep(3)  # Ожидание загрузки страницы товара

            # Парсим характеристики товара
            for col_index, cell in enumerate(row[1:], start=2):
                characteristic_name = sheet.cell(row=1, column=col_index).value  # Название характеристики

                if not characteristic_name:
                    continue

                try:
                    # Найдите селектор для характеристики на странице товара
                    characteristic_value = driver.find_element(By.XPATH, f"//*[contains(text(), '{characteristic_name}')]/following-sibling::*").text
                    cell.value = characteristic_value
                except Exception as e:
                    print(f"Характеристика '{characteristic_name}' не найдена: {e}")

        except Exception as e:
            print(f"Товар '{product_name}' не найден: {e}")

finally:
    # Сохраняем изменения в Excel
    workbook.save(excel_path)
    print(f"Данные сохранены в {excel_path}")

    # Закрываем браузер
    driver.quit()
