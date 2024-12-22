import openpyxl
import openai
from config import api_key
from openai import OpenAI


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)

# Функция для получения описания от ChatGPT
def get_description(brand, collection):
    message = 'Напиши привлекательное описание для коллекции ' + collection + ' фабрики ' + brand
    content = 'Вы креативный копирайтер. Найди в интернете технические характеристики и фото этого товара.Проанализируй фото и информацию о товаре, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. Твое описание товара должно быть яркими и привлекать внимание. Описание должно быть на 100% уникальное!'
    try:
        chat_completion = client.chat.completions.create(
            model="chatgpt-4o-latest",
            max_tokens=500,
            temperature=0.7,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        print(f"Ошибка при обработке запроса для {brand} {collection}: {e}")
        return "Ошибка при генерации описания"


# test = get_description('Керамин', 'Гламур')
# print(test)



# Открываем Excel-файл
excel_path = "../products.xlsx"
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# столбцы: A (Бренд), B (Коллекция), C (Описание)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
    brand = row[0].value  # Значение из столбца "Бренд"
    collection = row[1].value  # Значение из столбца "Коллекция"

    if brand and collection:
        print(f"Обрабатываю: Бренд - {brand}, Коллекция - {collection}")
        description = get_description(brand, collection)
        row[2].value = description  # Записываем описание в столбец "Описание"

# Сохраняем изменения в Excel-файле
workbook.save(excel_path)
print(f"Описание сохранено.")
