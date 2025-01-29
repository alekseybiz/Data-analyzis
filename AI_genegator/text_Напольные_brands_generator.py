import openpyxl
import openai
from config import api_key
from openai import OpenAI


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

excel_path = "Напольные_brands.xlsx"

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)

# Функция для получения описания от ChatGPT
def get_description(brand, country):
    message = 'Напиши привлекательное описание для фабрики по производству напольных покрытий (ламинат, паркетная доска, кварцвиниловая плитка, инженерная доска и т.п.) ' + brand + ', расположенной в стране: ' + country
    print(message)
    content = 'Ты креативный копирайтер. Сделай описание фабрики: опиши ее особенности, особенности выпускаемой продукции. Можно вставить какие-то интересные факты о фабрике. Описание должно быть ярким и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Пиши от третьего лица. Используй html-разметку, но не вставляй в описание "```html". Каждый абзац должен быть в каком-то html-теге. Все парные html-теги обязательно должны быть обязательно закрыты - проверь это в полученном тексте! Описание должно быть на 100% уникальное! Описание не должно быть заспамлено ключевыми словами! ' + brand + ' (название фабрики) - можно использовать в тексте не более 5 раз на оригинальном языке + можно транслитерировать бренд на русский язык и написать его еще не более трех раз. Не вставляй ссылки на сайты!'
    print(content)
    try:
        chat_completion = client.chat.completions.create(
            model="chatgpt-4o-latest",
            max_tokens=1000,
            temperature=0.7,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        raise RuntimeError(f"Ошибка при обработке запроса для {brand}: {e}")

# Открываем Excel-файл
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active
i = 0

try:
    # Проход по строкам
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        brand = row[2].value  # Значение из столбца "Бренд"
        # if isinstance(brand, str) and brand:
        #     brand = brand.capitalize()
        # else:
        #     brand = ""
        country = row[1].value  # Значение из столбца "Страна"
        # if isinstance(country, str) and country:
        #     country = country.capitalize()
        # else:
        #     country = ""

        if brand:
            print(f">>>{i}. Обрабатываю фабрику - {brand}")
            try:
                description = get_description(brand, country)
            except RuntimeError as e:
                print(e)
                break  # Прерываем цикл при ошибке
            row[3].value = description  # Записываем описание в столбец "Описание"
        i += 1
        if i % 20 == 0:
            workbook.save(excel_path)
            print(f">> Описание сохранено в {excel_path}.")

finally:
    # Сохраняем изменения в Excel-файле независимо от результата
    workbook.save(excel_path)
    print(f"Описание сохранено в {excel_path}.")
