import openpyxl
import openai
from config import api_key
from openai import OpenAI
from openpyxl.styles import Alignment


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)


# Функция для генерации описания товара через GPT
def get_description(product_info):
    prompt = (
        f"Составь краткое, привлекательное и убедительное описание для товара на основе его характеристик:\n"
        f"{product_info}\n\n"
        f"Описание должно быть на 100% уникальное!"
        # f"Проанализируй фото и информацию о товаре, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. \n"
    )
    print(prompt)
    try:
        response = client.chat.completions.create(
            model="chatgpt-4o-latest",
            messages=[
                {"role": "system", "content": "Ты профессиональный креативный копирайтер, создающий описания товаров для маркетплейса OZON."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=600,
            temperature=0.6
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Ошибка при генерации описания: {e}")
        return "Ошибка генерации описания"


# Открываем Excel-файл
excel_path = "products_2_0.xlsx"
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# Добавляем заголовок для колонки "Описание товара" в конец листа
if "Описание товара" not in [cell.value for cell in sheet[1]]:
    sheet.cell(row=1, column=sheet.max_column + 1, value="Описание товара")

# Определяем индекс колонки для "Описание товара"
description_col_index = sheet.max_column

# Обрабатываем строки с товарами
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    product_name = row[0].value  # Название товара из первой колонки
    product_info = f"Товар: {product_name}"  # Начинаем строку с названия товара

    # Сбор характеристик из остальных колонок
    for col_index, cell in enumerate(row[1:]):  # Начинаем со второй колонки
        header = sheet.cell(row=1, column=col_index + 2).value  # Название свойства из заголовка
        value = cell.value
        if value:  # Пропускаем пустые ячейки
            product_info += f"; {header}: {value}"

    print(f"product_info: {product_info}")

    # Генерируем описание товара
    description = get_description(product_info)
    print(f"description: {description}")

    # Записываем описание в соответствующую колонку
    # row[description_col_index - 1].value = description
    cell = row[description_col_index - 1]
    cell.value = description
    cell.alignment = Alignment(wrap_text=True)

# Сохраняем изменения в Excel-файле
workbook.save(excel_path)
print(f"Все описания сохранены в файле {excel_path}.")

