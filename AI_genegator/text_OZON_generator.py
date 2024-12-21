import openpyxl
import openai
from config import api_key
from openai import OpenAI


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

excel_path = "products_3_try.xlsx"

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)

# Функция для получения описания от ChatGPT
def get_description(product, brand, collection):
    message = 'Напиши небольшое привлекательное описание для товара ' + product + ' фабрики ' + brand + ' коллекции ' + collection
    # print(message)
    content = 'Вы креативный копирайтер. Найди в интернете технические характеристики и фото этого товара. Проанализируй фото и информацию о товаре, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. Твое описание товара должно быть яркими и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Не используй разметку Markdown (пиши без символов "*"). Описание должно быть на 100% уникальное! '
    try:
        chat_completion = client.chat.completions.create(
            model="chatgpt-4o-latest",
            max_tokens=600,
            temperature=0.7,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        raise RuntimeError(f"Ошибка при обработке запроса для {product}: {e}")

        # print(f"Ошибка при обработке запроса для {product}: {e}")
        # return "Ошибка при генерации описания"


# Открываем Excel-файл
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

try:
    # Проход по строкам
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        product = row[2].value  # Значение из столбца "товар"
        brand = row[3].value  # Значение из столбца "Бренд"
        if isinstance(brand, str) and brand:
            brand = brand.capitalize()
        else:
            brand = ""
        collection = row[4].value  # Значение из столбца "Коллекция"
        if isinstance(collection, str) and collection:
            collection = collection.capitalize()
        else:
            collection = ""

        if product:
            print(f"Обрабатываю: товар - {product}")
            try:
                description = get_description(product, brand, collection)
            except RuntimeError as e:
                print(e)
                break  # Прерываем цикл при ошибке
            row[5].value = description  # Записываем описание в столбец "Описание"

finally:
    # Сохраняем изменения в Excel-файле независимо от результата
    workbook.save(excel_path)
    print(f"Описание сохранено в {excel_path}.")
