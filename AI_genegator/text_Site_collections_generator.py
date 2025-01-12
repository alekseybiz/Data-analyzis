import openpyxl
import openai
from config import api_key
from openai import OpenAI


# Установите ваш API-ключ OpenAI
openai.api_key = api_key

excel_path = "тексты коллекций_try.xlsx"

client = OpenAI(
    api_key=api_key,  # This is the default and can be omitted
)

# Функция для получения описания от ChatGPT
def get_description(brand, collection):
    message = 'Напиши небольшое привлекательное описание для коллекции плитки ' + collection + ' фабрики ' + brand
    print(message)
    content = 'Ты креативный копирайтер. Найди в интернете технические характеристики и фото этой коллекции. Проанализируй фото и информацию о плитках в этой коллекции, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. Твое описание товара должно быть яркими и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Пиши от третьего лица. Используй html-разметку, но не вставляй в описание "```html". Каждый абзац должен быть в каком-то html-теге. Все парные html-теги обязательно должны быть обязательно закрыты - проверь это в полученном тексте! Описание должно быть на 100% уникальное! Описание не должно быть заспамлено ключевыми словами! ' + collection + ' (название коллекции) - можно использовать в тексте не более 4 раз. Не вставляй ссылки на сайты!'
    # content = 'Ты креативный копирайтер. Сделай описание фабрики: опиши ее особенности, особенности выпускаемой продукции. Можно вставить какие-то интересные факты о фабрике. Описание должно быть ярким и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Пиши от третьего лица. Используй html-разметку, но не вставляй в описание "```html". Каждый абзац должен быть в каком-то html-теге. Все парные html-теги обязательно должны быть обязательно закрыты - проверь это в полученном тексте! Описание должно быть на 100% уникальное! Описание не должно быть заспамлено ключевыми словами! ' + brand + ' (название фабрики) - можно использовать в тексте не более 4 раз на оригинальном языке + можно транслитерировать бренд на русский язык и написать его еще не более трех раз. Не вставляй ссылки на сайты!'

    try:
        chat_completion = client.chat.completions.create(
            # model="gpt-4o-mini",
            model="chatgpt-4o-latest",
            max_tokens=900,
            temperature=0.7,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        raise RuntimeError(f"Ошибка при обработке запроса для {collection}: {e}")

# Открываем Excel-файл
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active
i = 0

try:
    # Проход по строкам
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        brand = row[2].value  # Значение из столбца "Бренд"
        if isinstance(brand, str) and brand:
            brand = brand.capitalize()
        else:
            brand = ""
        collection = row[3].value  # Значение из столбца "Коллекция"
        if isinstance(collection, str) and collection:
            collection = collection.capitalize()
        else:
            collection = ""
        print(f"Бренд: {brand}, коллекция: {collection}")
        if collection:
            print(f"{i}. Обрабатываю коллекцию: {collection}")
            try:
                while True:  # Цикл для проверки и генерации текста
                    description = get_description(brand, collection)
                    if description.endswith(">"):  # Проверяем последний символ
                        break  # Если проверка пройдена, выходим из цикла
                    print("Описание не прошло проверку, повторяем генерацию.")
                print(f"Описание: {description}")
            except RuntimeError as e:
                print(e)
                break  # Прерываем цикл при ошибке
            row[4].value = description  # Записываем описание в столбец "Описание"
        i += 1

finally:
    # Сохраняем изменения в Excel-файле независимо от результата
    workbook.save(excel_path)
    print(f"Описание сохранено в {excel_path}.")
