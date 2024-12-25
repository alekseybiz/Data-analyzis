import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains



# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "for_parsing_try.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# URL сайта для поиска
base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта


# Проходим по всем строкам таблицы
for row in sheet.iter_rows(min_row=42, max_row=sheet.max_row):
    brand = row[4].value  # Значение из 5-го столбца ("Brand")
    brand = brand.split()[0] if " " in brand else brand
    print(brand)  # Вывод первого слова в Бренд
    collection = row[5].value  # Значение из 6-го столбца ("Collection")

    # Проверяем, что значения не пустые
    if not brand and not collection:
        continue

    brand_and_collection = brand + " " + collection
    # brand_and_collection = "Realistik Ginza"
    print(f"Ищем: {brand_and_collection}")

    search_url = f"https://3dplitka.ru/search/?q={brand_and_collection.replace(' ', '%20')}"
    driver.get(search_url)
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов ???

    results = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-container"))
    )
    elements_in_collection = len(results)
    print(f"Найдено элементов в коллекции: {elements_in_collection}")

    # Проходим по всем элементам коллекции
    for index in range(elements_in_collection):
        try:
            # Повторно загружаем элементы коллекции, чтобы избежать устаревания ссылок
            results = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-container a"))
            )

            # Берём текущий элемент по индексу
            product_link = results[index]
            product_href = product_link.get_attribute("href")
            print(f"Открываем элемент {index + 1}: {product_href}")
            #
            # # Переходим по ссылке
            # driver.get(product_href)
            #
            # # Проверяем текущий URL
            # current_url = driver.current_url
            # print(f"Открыт URL товара: {current_url}")
            #
            # # Возвращаемся обратно к коллекции
            # driver.get(search_url)
            # time.sleep(3)  # Ожидание загрузки страницы коллекции
        except Exception as e:
            print(f"Ошибка при обработке элемента {index + 1}: {e}")

    # # Открываем первую ссылку на найденный товар
    # product_link = driver.find_element(By.CSS_SELECTOR, ".product-card-container a")
    # product_href = product_link.get_attribute("href")
    # # Переходим по извлечённой ссылке
    # driver.get(product_href)
    #
    # # Проверяем текущий URL
    # current_url = driver.current_url
    # print(f"Открыт URL товара: {current_url}")







# Сохраняем изменения в Excel
workbook.save(excel_path)
print(f"Данные сохранены в {excel_path}")

# Закрываем браузер
driver.quit()
