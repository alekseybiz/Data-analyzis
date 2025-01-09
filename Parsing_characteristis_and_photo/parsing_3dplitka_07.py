# https://tile.expert/
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import urllib.request
# from P_S_engine_google import search_images, download_image, contains_watermark, save_image  # Импортируем функции
# import os
# import requests
# from PIL import Image
# from io import BytesIO
# import hashlib
# from config import console_cloud_google_API_1, search_engine_id
# import pytesseract
#
# # Укажите путь к Tesseract
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Путь к tesseract
#
# API_KEY = console_cloud_google_API_1
# CX = search_engine_id
#
# MIN_SIZE = 700
# # MAX_RETRIES = 3
# image_hashes = set()
# saved_images_count = 0

#
# # Путь для сохранения изображений
# SAVE_FOLDER = "downloaded_images"
# if not os.path.exists(SAVE_FOLDER):
#     os.makedirs(SAVE_FOLDER)

# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "for_parsing_try.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# # URL сайта для поиска
# base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта


# Проходим по всем строкам таблицы
row_number = 43  # Начальный номер строки
while row_number <= sheet.max_row:
    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    brand = row[4].value  # Значение из 5-го столбца ("Brand")
    if not brand:
        row_number += 1
        continue
    brand = brand.split()[0] if " " in brand else brand
    print(f"Новая строка, brand= {brand}")  # Вывод первого слова в Бренд
    collection = row[5].value  # Значение из 6-го столбца ("Collection")
    if not collection:
        row_number += 1
        continue
    # Проверяем, что значения не пустые
    if not brand and not collection:
        row_number += 1
        continue

    brand_and_collection = brand + " " + collection
    print(f"Ищем: {brand_and_collection}")

    search_url = f"https://3dplitka.ru/search/?q={brand_and_collection.replace(' ', '%20')}"
    driver.get(search_url)
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов

    results = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-container"))
    )
    elements_in_collection = len(results)
    print(f"Найдено элементов в коллекции: {elements_in_collection}")

    # Проходим по всем элементам в коллекции
    for index in range(elements_in_collection):
        # Повторно загружаем элементы коллекции, чтобы избежать устаревания ссылок
        results = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.product-card.narrow a")))
        product_link = results[index]
        product_href = product_link.get_attribute("href")
        driver.get(product_href)
        current_url = driver.current_url
        print(f"Открыт URL товара: {current_url}")
        # Проверяем есть в url 'product':
        if "product" not in current_url:
            print(f"Слова 'product' нет в этом url.")
            continue

        # Нажимаем Кнопку 'Показать всё'.
        try:
            button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div.vue-foldable-view-more.collapsed")))
            button.click()
            print("Кнопка 'Показать всё' нажата.")
        except Exception as e:
            print(f"Произошла ошибка. Нет кнопки 'Показать все' ")
            # Если ошибка - следующий
            continue


        # 1. Наименование товара (кол. №7)
        col_number = 7
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.el-col.el-col-8 h1")))
        # Извлечение текста элемента
        product_name = element.text.replace(" - керамическая плитка и керамогранит", "")
        print(f"Название элемента: {product_name}")

        # Проверяем есть в названии Бренд и Коллекция:
        if brand.lower() not in product_name.lower() and collection.lower() not in product_name.lower():
            print(f"Бренда {brand} и Коллекции {collection} нет в этом товаре.")
            continue

        if index > 0:
            # Вставляем новую строку ниже текущей строки
            sheet.insert_rows(row_number + 1)
            row_number += 1  # Обновляем текущий номер строки после вставки
            # Заполняем столбцы №№ 2...6
            for i in range(2, 7):
                sheet.cell(row=row_number, column=i).value = sheet.cell(row=row_number - 1, column=i).value



        # ПАРСИНГ
        # 1. Записываем колонку №1:
        sheet.cell(row=row_number, column=1).value = current_url

        # 2. Записываем Название товара:
        sheet.cell(row=row_number, column=col_number).value = product_name

        # 3. Все характеристики
        properties = ['Назначение', 'Материал', 'Основной цвет', 'Цветовые оттенки',
                      'Отражение поверхности', 'Обработка', 'Имитация', 'Стиль', 'Форма', 'Количество Лиц',
                      'Вариативность цвета', 'Морозоустойчивость', 'Противоскользящая', 'Сопротивление скольжению',
                      'Износостойкость', 'Влагопоглощаемость', 'В упаковке', 'Кол-во м2 в упаковке',
                      'Ширина, см', 'Длина, см', 'Толщина мм', 'Вес 1 шт.', 'Вес упаковки']
        col_number = 9
        for prop in properties:
            # Находим все контейнеры с классом attr-row divided
            containers = WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.attr-row.divided")))
            # Перебираем все найденные контейнеры
            for container in containers:
                try:
                    # Извлекаем текст из заголовка
                    header_element = container.find_element(By.CSS_SELECTOR, "span.attr-name")
                    header_text = header_element.text.strip().replace("\n", "")
                    # Проверяем, что заголовок равен prop:
                    if prop in header_text:
                        # Извлекаем текст значения
                        value_element = container.find_element(By.CSS_SELECTOR, "div.raw-content.attr-value")
                        text = value_element.text
                        # # Ищем и заменяем "Для коридора и кухни" на "Для коридора, Для кухни"
                        text = text.replace("Для коридора и кухни", "Для коридора, Для кухни")
                        text = text.replace(" шт", "")
                        text = text.replace(" м2", "")
                        text = text.replace(" кг", "")
                        text = text.replace(" /", ",")

                        print(f"Значение для '{header_text}': {text}")
                        sheet.cell(row=row_number, column=col_number).value = text
                        # Записываем заголовок:
                        if not sheet.cell(row=row_number - 1, column=col_number).value:
                            sheet.cell(row=row_number - 1, column=col_number).value = header_text
                        break  # Прерываем цикл, если нашли нужный контейнер
                except Exception as e:
                    print(f"Ошибка при обработке контейнера: {e}")
            else:
                print(f"Контейнер с заголовком {prop} не найден.")
            col_number += 1

        # 4. Цена товара (кол. №41)
        col_number = 41
        # Находим контейнер <p> с уникальным классом "card-price"
        container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "card-price")))
        # Внутри этого контейнера ищем <span class="card-attr-value">
        element = container.find_element(By.CSS_SELECTOR, "span.card-attr-value")
        product_price = element.text.strip().replace(" руб./м²", "")
        product_price = product_price.replace(" ", "")
        print(f"Цена элемента: {product_price}")
        sheet.cell(row=row_number, column=col_number).value = product_price

        # 5. Тип товара (кол. №8)
        col_number = 8
        container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".product-attrs-block")))
        rows = container.find_elements(By.CSS_SELECTOR, ".attr-row")
        product_type = None
        for row in rows:
            # Ищем элемент с классом "attr-name"
            attr_name = row.find_element(By.CSS_SELECTOR, ".attr-name").text.strip()
            if attr_name == "Тип товара":
                # Если нашли, берём значение из элемента с классом "attr-value"
                product_type = row.find_element(By.CSS_SELECTOR, ".attr-value").text.strip()
                break
        if product_type:
            print(f"Тип товара: {product_type}")
            sheet.cell(row=row_number, column=col_number).value = product_type
        else:
            print("Тип товара не найден.")

        # 6. Фото товара (кол. №44)
        col_number = 44



    row_number += 1


# Сохраняем изменения в Excel
workbook.save(excel_path)
print(f"Данные сохранены в {excel_path}")

# Закрываем браузер
driver.quit()
