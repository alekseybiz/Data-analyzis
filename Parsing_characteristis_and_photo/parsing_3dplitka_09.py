import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "parsing_3dplitka/Керам-Трейд 12.12.24.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# # URL сайта для поиска
# base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта


# Проходим по всем строкам таблицы
row_number = 43  # Начальный номер строки
while row_number <= sheet.max_row:
    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    brand = row[5].value  # Значение из 6-го столбца ("Brand")
    brand_ours = brand
    three_dplitka_brand = row[1].value  # Значение из 2-го столбца (Название фабрики в 3dplitka)
    if not brand and not three_dplitka_brand:
        row_number += 1
        continue
    # brand = brand.split()[0] if " " in brand else brand  # Вывод первого слова в Бренд
    if three_dplitka_brand:
        brand = three_dplitka_brand

    print(f"Стр.{row_number}. brand: {brand}")
    collection = row[6].value  # Значение из 7-го столбца ("Collection")
    if not collection:
        row_number += 1
        continue
    # # Проверяем, что значения не пустые
    # if not brand and not collection:
    #     row_number += 1
    #     continue

    brand_and_collection = brand + " " + collection
    print(f"Ищем: {brand_and_collection}")

    search_url = f"https://3dplitka.ru/search/?q={brand_and_collection.replace(' ', '%20')}"
    driver.get(search_url)
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов

    # Нажимаем Кнопку 'Показать еще N товаров'.
    try:
        button = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.show-all-btn")))
        button.click()
        print("Кнопка 'Показать еще N товаров' нажата.")
    except Exception as e:
        print(f"Нет кнопки 'Показать еще N товаров' ")

    # Прокручиваем страницу вниз, чтобы все элементы стали видимыми
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(3)  # Ожидание отображения всех элементов после прокрутки

    try:
        results = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-image")))
    except Exception as e:
        print(f"! results не найден. Ошибка {e}")
        row_number += 1
        continue

    elements_in_collection = len(results)
    print(f"Найдено элементов в коллекции: {elements_in_collection}")

    # Собираем все ссылки на элементы товаров
    product_links = set()  # Используем множество для хранения уникальных ссылок
    for index, result in enumerate(results):
        try:
            # Находим <a> внутри карточки товара
            link_element = result.find_element(By.XPATH,
                                               ".//ancestor::div[contains(@class, 'product-card-container')]//a")
            product_href = link_element.get_attribute("href")  # Получаем ссылку на товар
            if product_href not in product_links:  # Проверяем, есть ли уже такая ссылка
                product_links.add(product_href)  # Добавляем ссылку в множество
                print(f"{index + 1}. Уникальная ссылка на товар: {product_href}")
        except Exception as e:
            print(f"Ошибка при извлечении ссылки из элемента {index}: {e}")
    # Преобразуем множество обратно в список, если это нужно
    product_links = list(product_links)

        #     product_links.append(product_href)  # Сохраняем ссылку
        #     print(f"{index+1}. Ссылка на товар: {product_href}")
        # except Exception as e:
        #     print(f"Ошибка при извлечении ссылки из элемента {index}: {e}")

    # Переходим по каждой ссылке
    for index, product_href in enumerate(product_links):
        driver.get(product_href)
        current_url = driver.current_url
        print(f"Открыт URL товара {index+1}: {current_url}")

        # Проверяем: есть ли в url 'product':
        if "product" not in current_url:
            print(f"! Слова 'product' нет в этом url.")
            continue

        # Нажимаем Кнопку 'Показать всё'.
        try:
            button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div.vue-foldable-view-more.collapsed")))
            button.click()
            print("Кнопка 'Показать всё' нажата.")
        except Exception as e:
            print(f"Нет кнопки 'Показать все' ")
            # Если ошибка - следующий
            # continue

        # 1. Наименование товара (кол. №7)
        col_number = 8
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.el-col.el-col-8 h1")))
        # Извлечение текста элемента
        product_name = element.text.replace(" - керамическая плитка и керамогранит", "")
        print(f"Название элемента: {product_name}")

        # Проверяем есть ли в названии Бренд и Коллекция:
        search_words = brand_and_collection.lower().split()
        # Проверяем наличие всех слов в названии товара
        if not all(word in product_name.lower() for word in search_words):
            print(f"! Не все слова из '{brand_and_collection}' найдены в названии товара: '{product_name}'")
            continue

        if index > 0:
            # Вставляем новую строку ниже текущей строки
            sheet.insert_rows(row_number + 1)
            row_number += 1  # Обновляем текущий номер строки после вставки
            # Заполняем столбцы №№ 2...8
            for i in range(2, 8):
                sheet.cell(row=row_number, column=i).value = sheet.cell(row=row_number - 1, column=i).value


        # ПАРСИНГ
        # 1. Записываем колонку №1:
        sheet.cell(row=row_number, column=1).value = current_url

        # 2. Записываем Название товара:
        sheet.cell(row=row_number, column=col_number).value = product_name

        # 3. Все характеристики
        properties = ['Назначение', 'Материал', 'Основной цвет', 'Цветовые оттенки',
                      'Отражение поверхности', 'Обработка', 'Имитация', 'Стиль', 'Форма', 'Количество Лиц',
                      'Вариативность цвета', 'Морозоустойчивость', 'Противоскользящая', 'Сопротивление скольжению',
                      'Износостойкость', 'Влагопоглощаемость', 'В упаковке', 'Кол-во м2 в упаковке',
                      'Ширина, см', 'Длина, см', 'Толщина мм', 'Вес 1 шт.', 'Вес упаковки']
        col_number = 10
        for prop in properties:
            # Находим все контейнеры с классом attr-row divided
            containers = WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.attr-row.divided")))
            # Перебираем все найденные контейнеры
            for container in containers:
                try:
                    # Извлекаем текст из заголовка
                    header_element = container.find_element(By.CSS_SELECTOR, "span.attr-name")
                    header_text = header_element.text.strip().replace("\n", "")
                    # Проверяем, что заголовок равен prop:
                    if prop in header_text:
                        # Извлекаем текст значения
                        value_element = container.find_element(By.CSS_SELECTOR, "div.raw-content.attr-value")
                        text = value_element.text
                        # # Ищем и заменяем "Для коридора и кухни" на "Для коридора, Для кухни"
                        text = text.replace("Для коридора и кухни", "Для коридора, Для кухни")
                        text = text.replace(" шт", "")
                        text = text.replace(" м2", "")
                        text = text.replace(" кг", "")
                        text = text.replace(" /", ",")

                        print(f"Значение для '{header_text}': {text}")
                        sheet.cell(row=row_number, column=col_number).value = text
                        # Записываем заголовок:
                        if not sheet.cell(row=row_number - 1, column=col_number).value:
                            sheet.cell(row=row_number - 1, column=col_number).value = header_text
                        break  # Прерываем цикл, если нашли нужный контейнер
                except Exception as e:
                    print(f"Ошибка при обработке контейнера: {e}")
            else:
                print(f"! Контейнер с заголовком {prop} не найден.")
            col_number += 1

        # 4. Цена на 3dplitka (кол. №42)
        col_number = 42
        # Находим контейнер <p> с уникальным классом "card-price"
        container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "card-price")))
        # Внутри этого контейнера ищем <span class="card-attr-value">
        element = container.find_element(By.CSS_SELECTOR, "span.card-attr-value")
        product_price = element.text.strip().replace(" руб./м²", "")
        product_price = product_price.replace(" ", "")
        print(f"Цена элемента: {product_price}")
        sheet.cell(row=row_number, column=col_number).value = product_price

        # 5. Тип товара (кол. №9)
        col_number = 9
        container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".product-attrs-block")))
        rows = container.find_elements(By.CSS_SELECTOR, ".attr-row")
        product_type = None
        for row in rows:
            # Ищем элемент с классом "attr-name"
            attr_name = row.find_element(By.CSS_SELECTOR, ".attr-name").text.strip()
            if attr_name == "Тип товара":
                # Если нашли, берём значение из элемента с классом "attr-value"
                product_type = row.find_element(By.CSS_SELECTOR, ".attr-value").text.strip()
                break
        if product_type:
            print(f"Тип товара: {product_type}")
            sheet.cell(row=row_number, column=col_number).value = product_type
        else:
            print("! Тип товара не найден.")

        # 6. Дополняем Наименование товара (кол. №8)
        col_number = 8
        product_name = product_name.replace(collection, f"{collection} {product_type}")
        if three_dplitka_brand:
            product_name = product_name.replace(three_dplitka_brand, brand_ours)

        sheet.cell(row=row_number, column=col_number).value = product_name
        print(f"Название элемента дополненное: {product_name}")


        # 6. Фото товара (кол. №44)
        # col_number = 44


    # Сохраняем изменения в Excel
    workbook.save(excel_path)
    print(f"Данные стр. {row_number} сохранены в {excel_path}")
    row_number += 1
    # print(f"row_number: {row_number}")


# Закрываем браузер
driver.quit()
