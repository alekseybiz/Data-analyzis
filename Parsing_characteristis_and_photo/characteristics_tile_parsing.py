import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains



# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "products.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# URL сайта для поиска
base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта

try:
    # Проходим по всем строкам таблицы
    for row in sheet.iter_rows(min_row=42, max_row=sheet.max_row):
        brand = row[4].value  # Значение из 5-го столбца ("Brand")
        collection = row[5].value  # Значение из 6-го столбца ("Collection")

        # Проверяем, что значения не пустые
        if not brand and not collection:
            continue

        brand_and_collection = brand + " " + collection
        # brand_and_collection = "Realistik Ginza"
        print(f"Ищем: {brand_and_collection}")

        search_url = f"https://3dplitka.ru/search/?q={brand_and_collection.replace(' ', '%20')}"
        driver.get(search_url)
        current_url = driver.current_url
        print(f"Открыт URL поиска: {current_url}")
        time.sleep(3)  # Ожидание загрузки результатов


        try:
            results = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-container"))
            )
            print(f"Найдено элементов коллекции: {len(results)}")
        except Exception as e:
            print(f"Ошибка: {e}")


        try:
            # Открываем первую ссылку на найденный товар
            product_link = driver.find_element(By.CSS_SELECTOR, ".product-card-container a")
            product_href = product_link.get_attribute("href")
            # Переходим по извлечённой ссылке
            driver.get(product_href)

            # Проверяем текущий URL
            current_url = driver.current_url
            print(f"Открыт URL товара: {current_url}")





            time.sleep(3)  # Ожидание загрузки страницы товара
        #
        #     # Парсим характеристики товара
        #     for col_index, cell in enumerate(row[1:], start=2):
        #         characteristic_name = sheet.cell(row=1, column=col_index).value  # Название характеристики
        #
        #         if not characteristic_name:
        #             continue
        #
        #         try:
        #             # Найдите селектор для характеристики на странице товара
        #             characteristic_value = driver.find_element(By.XPATH, f"//*[contains(text(), '{characteristic_name}')]/following-sibling::*").text
        #             cell.value = characteristic_value
        #         except Exception as e:
        #             print(f"Характеристика '{characteristic_name}' не найдена: {e}")

        except Exception as e:
            print(f"Товар '{product_name}' не найден: {e}")

finally:
    # Сохраняем изменения в Excel
    workbook.save(excel_path)
    print(f"Данные сохранены в {excel_path}")

    # Закрываем браузер
    driver.quit()
