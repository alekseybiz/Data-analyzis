import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import openai
from config import api_key
from openai import OpenAI

# Установите ваш API-ключ OpenAI
openai.api_key = api_key
client = OpenAI(api_key=api_key)  # This is the default and can be omitted

# Настройте путь к вашему файлу Excel
excel_path = "Ozon/Озон ИП 2025-01-22.xlsx"

# Функция для нахождения характеристик от ChatGPT
def get_description(product_name, property, variants, explanation):
    message = 'Для плитки ' + product_name + " найди свойство: " + property
    print(f"message: {message}")
    # content = 'Ты креативный копирайтер. Найди в интернете технические характеристики и фото этой коллекции. Проанализируй фото и информацию о плитках в этой коллекции, выдели ключевые характеристики и преимущества. Проанализируй стиль, цвет, форму, поверхность, размеры товара, его художественное исполнение. Твое описание товара должно быть яркими и привлекать внимание. Добавь уникальности, пиши как топовый маркетолог, внеси в текст интересные факты. Пиши от третьего лица. Используй html-разметку, но не вставляй в описание "```html". Каждый абзац должен быть в каком-то html-теге. Все парные html-теги обязательно должны быть обязательно закрыты - проверь это в полученном тексте! Описание должно быть на 100% уникальное! Описание не должно быть заспамлено ключевыми словами! ' + collection + ' (название коллекции) - можно использовать в тексте не более 4 раз. Не вставляй ссылки на сайты!'
    # content = 'Ты должен дать четкий ответ на вопрос. Вот пояснения от ' + explanation + '. Ответ предоставь в соответствии с пояснениями OZON: если допускается множественное значение, то перечисли значения используя разделитель "; ", выбери значения строго из списка: ' + variants + '! А если нужно только одно значение, то выдай только одно значение. Если ' + variants + ' пусто, значит ответ может быть произвольным. Ответ должен содержать ТОЛЬКО итоговое значение или значения (если допускается множественный выбор)! Не пиши никаких комментариев, нужно ТОЛЬКО значение или значения!'
    content = 'Ты должен дать четкий ответ на вопрос. Вот пояснения от ' + (
                explanation or "нет пояснений") + '. Ответ предоставь в соответствии с пояснениями OZON: если допускается множественное значение, то перечисли значения используя разделитель "; ", выбери значения строго из списка: ' + (
                          variants or "нет вариантов") + '! А если нужно только одно значение, то выдай только одно значение. Если ' + (
                          variants or "нет вариантов") + ' пусто, значит ответ может быть произвольным. Ответ должен содержать ТОЛЬКО итоговое значение или значения (если допускается множественный выбор)! Не пиши никаких комментариев, нужно ТОЛЬКО значение или значения! Если ты не знаешь ответ, то в ответ не пиши НИЧЕГО - в ответе должна быть пустая строка!'

    print(f"content: {content}")

    try:
        chat_completion = client.chat.completions.create(
            model="gpt-4o-mini",
            # model="chatgpt-4o-latest",
            max_tokens=500,
            temperature=0.1,
            messages=[
                {
                    "role": "system",
                    "content": content,
                },
                {
                    "role": "user",
                    "content": message,
                }
            ]
        )
        # Извлекаем описание из ответа
        return chat_completion.choices[0].message.content.strip()
    except Exception as e:
        raise RuntimeError(f"Ошибка при обработке запроса для {product_name}: {e}")



# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

col_numbers = [9, 10, 11, 12, 19, 20, 21, 22, 23, 24, 25, 28, 31, 32, 33, 34, 35, 36, 40, 43, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
# col_numbers = [36]


# Проходим по всем строкам таблицы
row_number = 6  # Начальный номер строки
while row_number <= sheet.max_row:

    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    product_name = row[2].value  # Значение из 3-го столбца ("Название товара")
    if not product_name:
        row_number += 1
        continue
    print(f">>> Стр.{row_number}. товар: {product_name}")


    for col_number in col_numbers:
        property = sheet.cell(row=2, column=col_number).value
        print(f"property: {property}")
        variants = sheet.cell(row=4, column=col_number).value
        # print(f"variants: {variants}")
        explanation = sheet.cell(row=5, column=col_number).value
        # print(f"explanation: {explanation}")
        exist_property = sheet.cell(row=row_number, column=col_number).value
        new_property = ""
        if exist_property == "Пусто" or exist_property == "пусто":
            exist_property = ""
        print(f"exist_property: {exist_property}")

        if col_number in (24, 31, 33) and exist_property:
            # Проверяем, является ли exist_property строкой
            if isinstance(exist_property, str) and ',' in exist_property:
                new_property = exist_property.replace(',', '.')
                # Записываем свойство только если замена выполнена
                sheet.cell(row=row_number, column=col_number).value = new_property
                print(f"! Перезаписали new_property: {new_property}")
            elif not isinstance(exist_property, str):
                # Если это не строка, преобразуем в строку и проверяем наличие запятой
                exist_property_str = str(exist_property)
                if ',' in exist_property_str:
                    new_property = exist_property_str.replace(',', '.')
                    # Записываем свойство только если замена выполнена
                    sheet.cell(row=row_number, column=col_number).value = new_property
                    print(f"! Перезаписали new_property: {new_property}")

        if col_number == 41 and not exist_property:
            new_property = f"Цена за упаковку. В упаковке {sheet.cell(row=row_number, column=21).value} штук."
            sheet.cell(row=row_number, column=col_number).value = new_property
            print(f"! Перезаписали: {new_property}")
            # continue
        if col_number == 42 and not exist_property:
            new_property = 1
            sheet.cell(row=row_number, column=col_number).value = new_property
            print(f"! Перезаписали: {new_property}")


        if exist_property or new_property:
            continue

        description = get_description(product_name, property, variants, explanation)
        print(f"description сгенерировано: {description}")

        # Удаляем из description значения, которых нет в variants
        if variants:
            # Преобразуем строки в списки
            variants_list = [variant.strip() for variant in variants.split(";")]
            description_list = [desc.strip() for desc in description.split(";")]
            # Оставляем только те элементы из description, которые есть в variants
            filtered_description = [desc for desc in description_list if desc in variants_list]
            # Объединяем результат обратно в строку в description:
            description = "; ".join(filtered_description)
            print(f"description после чистки: {description}")

        # Записываем свойство:
        sheet.cell(row=row_number, column=col_number).value = description

    if row_number % 30 == 0:
        # print(f"{row_number} делится на 30 без остатка")
        workbook.save(excel_path)
        print(f"Данные стр. {row_number} сохранены в {excel_path}")

    row_number += 1

workbook.save(excel_path)
print(f"Данные стр. {row_number} сохранены в {excel_path}")





