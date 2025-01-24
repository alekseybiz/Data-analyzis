import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import requests
import os

# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "parsing_3dplitka/try.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# Создаем папку для сохранения изображений
os.makedirs("downloaded_images", exist_ok=True)

# Проходим по всем строкам таблицы
row_number = 44  # Начальный номер строки
while row_number <= sheet.max_row:
    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    brand = row[5].value  # Значение из 6-го столбца ("Brand")

    collection = row[6].value  # Значение из 7-го столбца ("Collection")
    if not collection:
        row_number += 1
        continue

    product_name = row[7].value  # Значение из 8-го столбца ("Наименование товара")
    print(f">>>Стр.{row_number}. товар: {product_name}")
    if not product_name:
        row_number += 1
        continue


    # 1. Ищем картинки на https://www.bestceramic.ru/
    search_url = f"https://www.bestceramic.ru/search?q={product_name.replace(' ', '+')}"
    driver.get(search_url)
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов


    try:
        result = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".item__head")))
    except Exception as e:
        print(f"! result не найден. Ошибка {e}")
        row_number += 1
        continue
    # Находим <a> внутри карточки товара
    link_element = result.find_element(By.CSS_SELECTOR, 'a[itemprop="url"]')
    link = link_element.get_attribute("href")  # Извлекаем ссылку
    print(f"Найдена ссылка: {link}")
    # Переходим по ссылке
    driver.get(link)
    print("Перешли по ссылке.")

    # Здесь можно выполнять действия на новой странице
    time.sleep(1)  # Ожидаем загрузки новой страницы


    image_number = 1

    # Находим первое изображение в карусели
    thumbnail = driver.find_element(By.CSS_SELECTOR, '.product-slider__item img')
    thumbnail.click()
    time.sleep(2)  # Ждем, пока откроется большое изображение

    # Находим первое большое изображение
    large_image = driver.find_element(By.CSS_SELECTOR, 'img.lg-object.lg-image')
    first_image_url = large_image.get_attribute("src")
    print(f"large_image_url: {first_image_url}")

    # Скачиваем первое изображение
    response = requests.get(first_image_url)
    with open(f"downloaded_images/{product_name}_{image_number}.jpg", "wb") as file:
        file.write(response.content)
    print(f"Скачано изображение: {first_image_url}")

    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    # Цикл для перехода по всем изображениям в карусели
    while True:
        try:
            # Находим кнопку "Next slide"
            next_button = driver.find_element(By.CSS_SELECTOR, "button.lg-next[aria-label='Next slide']")
            # Проверяем, активна ли кнопка
            if not next_button.is_displayed():
                print("Кнопка 'Next slide' не активна. Цикл завершён.")
                break

            # Нажимаем кнопку "Next slide"
            actions = ActionChains(driver)
            actions.move_to_element(next_button).click().perform()
            print(f"{image_number} Кнопка 'Next slide' нажата.")

            # Ждём изменения атрибута src у изображения
            try:
                WebDriverWait(driver, 10).until(
                    EC.staleness_of(large_image)  # Ожидаем, что элемент обновится (перестанет быть "старым")
                )
                time.sleep(2)  # Немного ждём для полной загрузки
                large_image = driver.find_element(By.CSS_SELECTOR, 'img.lg-object.lg-image')  # Находим заново
                large_image_url = large_image.get_attribute("src")
            except Exception as e:
                print(f"Ошибка ожидания изменения изображения: {e}")
                break

            print(f"large_image_url {image_number}: {large_image_url}")

            # Проверяем, если URL изображения совпадает с первым или предыдущим, завершаем цикл
            if large_image_url == first_image_url:
                print("Достигли первого изображения. Цикл завершён.")
                break

            # Увеличиваем счётчик изображений
            image_number += 1

            # Скачиваем новое изображение
            response = requests.get(large_image_url)
            with open(f"downloaded_images/{product_name}_{image_number}.jpg", "wb") as file:
                file.write(response.content)
            print(f"Скачано изображение: {large_image_url}")

            # Обновляем `first_image_url` для следующей проверки
            first_image_url = large_image_url

        except Exception as e:
            print(f"Ошибка при обработке карусели: {e}")
            break

 




    # Сохраняем изменения в Excel
    workbook.save(excel_path)
    print(f"Данные стр. {row_number} сохранены в {excel_path}")
    row_number += 1


# Закрываем браузер
driver.quit()
