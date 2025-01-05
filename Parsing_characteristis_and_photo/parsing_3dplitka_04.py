import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "for_parsing_try.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# URL сайта для поиска
base_url = "https://3dplitka.ru/"  # Замените на URL вашего сайта


# Проходим по всем строкам таблицы
row_number = 42  # Начальный номер строки
while row_number <= sheet.max_row:
    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    brand = row[4].value  # Значение из 5-го столбца ("Brand")
    brand = brand.split()[0] if " " in brand else brand
    print(f"Новая строка, brand= {brand}")  # Вывод первого слова в Бренд
    collection = row[5].value  # Значение из 6-го столбца ("Collection")
    print(f"row_number: {row_number}")
    # Проверяем, что значения не пустые
    if not brand and not collection:
        row_number += 1
        continue

    brand_and_collection = brand + " " + collection
    print(f"Ищем: {brand_and_collection}")

    search_url = f"https://3dplitka.ru/search/?q={brand_and_collection.replace(' ', '%20')}"
    driver.get(search_url)
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов

    results = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card-container"))
    )
    elements_in_collection = len(results)
    print(f"Найдено элементов в коллекции: {elements_in_collection}")

    # Проходим по всем элементам в коллекции
    for index in range(elements_in_collection):
        if index > 0:
            # Вставляем новую строку ниже текущей строки
            sheet.insert_rows(row_number + 1)
            row_number += 1  # Обновляем текущий номер строки после вставки
            print(f"row_number +1: {row_number}")

            for i in range(2, 7):  # Заполняем столбцы №№ 2...6
                sheet.cell(row=row_number, column=i).value = sheet.cell(row=row_number - 1, column=i).value

        # Повторно загружаем элементы коллекции, чтобы избежать устаревания ссылок
        results = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.product-card.narrow a"))
        )
        product_link = results[index]
        product_href = product_link.get_attribute("href")
        driver.get(product_href)
        current_url = driver.current_url
        print(f"Открыт URL товара: {current_url}")
        sheet.cell(row=row_number, column=1).value = current_url  # Записываем колонку №1

        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.vue-foldable-view-more.collapsed"))
        )
        button.click()
        print("Кнопка 'Показать всё' нажата.")


        # ПАРСИНГ
        # 1. Наименование товара (кол. №7)
        col_number = 7
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.el-col.el-col-8 h1")))
        # Извлечение текста элемента
        product_name = element.text.replace(" - керамическая плитка и керамогранит", "")
        print(f"Название элемента: {product_name}")
        sheet.cell(row=row_number, column=col_number).value = product_name

        # 2. Назначение (кол. №8)
        col_number = 8
        # Находим все контейнеры с классом attr-row divided
        containers = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.attr-row.divided"))
        )
        # Перебираем все найденные контейнеры
        for container in containers:
            try:
                # Извлекаем текст из заголовка
                header_element = container.find_element(By.CSS_SELECTOR, "span.attr-name")
                header_text = header_element.text

                # Проверяем, что заголовок равен "Назначение"
                if header_text == "Назначение":
                    # Извлекаем текст значения
                    value_element = container.find_element(By.CSS_SELECTOR, "div.raw-content.attr-value")
                    text = value_element.text.strip()
                    # Ищем и заменяем "Для коридора и кухни" на "Для коридора, Для кухни"
                    transformed_text = text.replace("Для коридора и кухни", "Для коридора, Для кухни")
                    print(f"Значение для '{header_text}': {transformed_text}")
                    sheet.cell(row=row_number, column=col_number).value = transformed_text
                    break  # Прерываем цикл, если нашли нужный контейнер
            except Exception as e:
                print(f"Ошибка при обработке контейнера: {e}")
        else:
            print("Контейнер с заголовком 'Назначение' не найден.")

        # 3. Цвет (кол. №9)
        col_number = 9
        # Находим все контейнеры с классом attr-row divided
        containers = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.attr-row.divided")))
        # Перебираем все найденные контейнеры
        for container in containers:
            try:
                # Извлекаем текст из заголовка
                header_element = container.find_element(By.CSS_SELECTOR, "span.attr-name")
                # header_text = header_element.get_attribute("innerText").strip()
                header_text = header_element.text.strip().replace("\n", "")
                # print(f"заголовок очищенный: {header_text}")

                # Проверяем, что заголовок равен "Основной цвет"
                # if header_text == "Основной цвет":
                if "Основной цвет" in header_text:
                    # Извлекаем текст значения
                    value_element = container.find_element(By.CSS_SELECTOR, "div.raw-content.attr-value")
                    text = value_element.text
                    # # Ищем и заменяем "Для коридора и кухни" на "Для коридора, Для кухни"
                    # transformed_text = text.replace("Для коридора и кухни", "Для коридора, Для кухни")
                    print(f"Значение для '{header_text}': {text}")
                    sheet.cell(row=row_number, column=col_number).value = text
                    break  # Прерываем цикл, если нашли нужный контейнер
            except Exception as e:
                print(f"Ошибка при обработке контейнера: {e}")
        else:
            print("Контейнер с заголовком 'Основной цвет' не найден.")


        # # Проходим по столбцам с 6 по 25
        # for col_index in range(7, 26):  # Нумерация столбцов начинается с 1
        #     head = sheet.cell(row=41, column=col_index).value
        #     cell_value = row[col_index - 1].value  #
        #     print(f"{head} : {cell_value}")

    row_number += 1


# Сохраняем изменения в Excel
workbook.save(excel_path)
print(f"Данные сохранены в {excel_path}")

# Закрываем браузер
driver.quit()
