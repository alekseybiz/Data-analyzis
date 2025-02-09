import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import requests
import os

# Настройте путь к вашему WebDriver
webdriver_path = r"C:\Users\Administrator\Documents\install\chromedriver\chromedriver.exe"  # Замените на путь к вашему драйверу

# Настройка сервиса для драйвера Chrome
service = Service(webdriver_path)

# Настройте путь к вашему файлу Excel
excel_path = "../parsing_3dplitka/try.xlsx"

# Инициализация браузера
driver = webdriver.Chrome(service=service)

# Открываем таблицу Excel
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# Создаем папку для сохранения изображений
os.makedirs("../downloaded_images", exist_ok=True)

# Проходим по всем строкам таблицы
row_number = 44  # Начальный номер строки
while row_number <= sheet.max_row:
    row = list(sheet.iter_rows(min_row=row_number, max_row=row_number))[0]  # Получаем текущую строку
    brand = row[5].value  # Значение из 6-го столбца ("Brand")

    collection = row[6].value  # Значение из 7-го столбца ("Collection")
    if not collection:
        row_number += 1
        continue

    product_name = row[7].value  # Значение из 8-го столбца ("Наименование товара")
    print(f">>>Стр.{row_number}. товар: {product_name}, бренд: {brand}")
    if not product_name:
        row_number += 1
        continue


    # 1. Ищем картинки на https://www.bestceramic.ru/
    search_url = f"https://www.bestceramic.ru/search?q={product_name.replace(' ', '+')}"
    driver.get(search_url) # Открываем страницу поиска
    collection_url = driver.current_url
    print(f"Открыт URL поиска: {collection_url}")
    time.sleep(3)  # Ожидание загрузки результатов

    # Находим все блоки с классом "item item--product"
    product_blocks = driver.find_elements(By.CSS_SELECTOR, ".item.item--product")
    print(f"product_blocks: {product_blocks}")


    # Перебираем найденные блоки и извлекаем данные
    for block in product_blocks:
        try:
            # Извлекаем текст из <a class="item__title">
            title_element = block.find_element(By.CSS_SELECTOR, "a.item__title")
            title_text = title_element.text
            print(f"Название: {title_text}")

            # Извлекаем текст из <span class="markitem__content">
            markitem_element = block.find_element(By.CSS_SELECTOR, "span.markitem__content")
            brand_text = markitem_element.text
            print(f"Бренд: {brand_text}")

            # Находим <a> внутри карточки товара
            link_element = block.find_element(By.CSS_SELECTOR, 'a[itemprop="url"]')
            link = link_element.get_attribute("href")  # Извлекаем ссылку
            print(f"Найдена ссылка: {link}")

        except Exception as e:
            print(f"Ошибка при обработке блока: {e}")

        if brand_text.lower() == brand.lower() and collection.lower() in title_text.lower():
            print(f"Найдено совпадение: {brand_text} и в названии коллекции {title_text} есть {collection}")
            break


    # Переходим по ссылке
    driver.get(link)
    print("Перешли по ссылке.")


    # Извлечение всех превью изображений
    thumbnails = driver.find_elements(By.CSS_SELECTOR, ".product-slider__item img")
    print(f"thumbnails: {thumbnails}")


    # Сохранение уникальных URL-адресов изображений
    unique_images = set()

    image_number = 1  # Счётчик изображений
    for thumbnail in thumbnails:
        # Получаем URL уменьшенного изображения
        thumbnail_src = thumbnail.get_attribute("src")

        # Преобразуем в URL большого изображения
        large_image_src = thumbnail_src.replace("/resize_cache/iblock/", "/iblock/").replace("/86_86_2/", "/")

        # Проверяем, не скачивали ли уже это изображение
        if large_image_src in unique_images:
            print(f"Повтор изображения: {large_image_src}, пропускаем.")
            continue

        # Добавляем в множество уникальных URL
        unique_images.add(large_image_src)

        print(f"Большое изображение URL: {large_image_src}")

        # Скачиваем изображение
        response = requests.get(large_image_src)
        if response.status_code == 200:
            with open(f"downloaded_images/{product_name}_{image_number}.jpg", "wb") as file:
                file.write(response.content)
            sheet.cell(row=row_number, column=44+image_number).value = f"downloaded_images/{product_name}_{image_number}.jpg"
            print(f"Изображение {image_number} сохранено.")
            image_number += 1
        else:
            print(f"Не удалось скачать изображение {image_number}.")


    # Сохраняем изменения в Excel
    workbook.save(excel_path)
    print(f"Данные стр. {row_number} сохранены в {excel_path}")
    row_number += 1


# Закрываем браузер
driver.quit()
